import io
import os
import openpyxl
import pandas as pd

UPLOAD_DIR = "/app/uploads"


def parse_excel(file_bytes: bytes) -> str:
    """
    Auto-detect the attachment format and parse accordingly:
    - SAP Vendor Setup/Change form → col B label + col C value parser
    - Simple CSR table → pandas table parser
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

        if "SAP VENDOR SET UP OR CHANGE" in wb.sheetnames:
            ws = wb["SAP VENDOR SET UP OR CHANGE"]
            lines = []
            for row in ws.iter_rows(values_only=True):
                label = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                value = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                if label and value and value not in ("SELECT ONE",) and not value.startswith("="):
                    lines.append(f"{label}: {value}")
                hlabel = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                hvalue = str(row[6]).strip() if len(row) > 6 and row[6] else ""
                if hlabel and hvalue and not hvalue.startswith("="):
                    lines.append(f"{hlabel}: {hvalue}")
            return "\n".join(lines) if lines else "(no form data extracted)"
        else:
            df = pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl")
            return df.to_string(index=False)
    except Exception as e:
        return f"(failed to parse Excel: {e})"


def save_attachment(request_id: str, version: int, filename: str, file_bytes: bytes) -> str:
    """Save file to disk and return the stored file path."""
    folder = os.path.join(UPLOAD_DIR, request_id)
    os.makedirs(folder, exist_ok=True)
    file_path = os.path.join(folder, f"v{version}_{filename}")
    with open(file_path, "wb") as f:
        f.write(file_bytes)
    return file_path
