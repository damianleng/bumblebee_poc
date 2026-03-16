import io
import pandas as pd
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, File, Form, Request, UploadFile, HTTPException
from sqlalchemy.orm import Session
from slowapi import Limiter
from slowapi.util import get_remote_address

from database import get_db
from models import Request as RequestModel, RequestItem
from agent import run_agent

limiter = Limiter(key_func=get_remote_address)
router = APIRouter()


def parse_excel(file_bytes: bytes) -> str:
    """
    Auto-detect the attachment format and parse accordingly:
    - SAP Vendor Setup/Change form → col B label + col C value parser
    - Simple CSR table → pandas table parser
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

        if "SAP VENDOR SET UP OR CHANGE" in wb.sheetnames:
            # Vendor form: read col B (label) + col C (value), and col E + col G for header fields
            ws = wb["SAP VENDOR SET UP OR CHANGE"]
            lines = []
            for row in ws.iter_rows(values_only=True):
                # Main fields: col B label → col C value
                label = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                value = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                if label and value and value not in ("SELECT ONE",) and not value.startswith("="):
                    lines.append(f"{label}: {value}")
                # Header fields: col E label → col G value
                hlabel = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                hvalue = str(row[6]).strip() if len(row) > 6 and row[6] else ""
                if hlabel and hvalue and not hvalue.startswith("="):
                    lines.append(f"{hlabel}: {hvalue}")
            return "\n".join(lines) if lines else "(no form data extracted)"
        else:
            # CSR table: simple pandas parse
            import pandas as pd
            df = pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl")
            return df.to_string(index=False)
    except Exception as e:
        return f"(failed to parse Excel: {e})"


@router.post("/api/ingest")
@limiter.limit("10/minute")
async def ingest(
    request: Request,
    sender: str = Form(...),
    subject: str = Form(...),
    email_body: str = Form(...),
    attachment: UploadFile = File(None),
    db: Session = Depends(get_db),
):
    timestamp = datetime.now(timezone.utc).isoformat()

    excel_table = "(no attachment)"
    if attachment and attachment.filename:
        raw_bytes = await attachment.read()
        excel_table = parse_excel(raw_bytes)

    try:
        result = run_agent(
            sender=sender,
            subject=subject,
            timestamp=timestamp,
            email_body=email_body,
            excel_table=excel_table,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Agent error: {e}")

    req = RequestModel(
        sender=sender,
        subject=subject,
        request_type=result.get("request_type"),
        vendor_number=result.get("vendor_number"),
        vendor_name=result.get("vendor_name"),
        confidence=result.get("confidence"),
        classification_status=result.get("classification_status"),
        notes=result.get("notes"),
        raw_agent_output=result,
        status="pending_review",
    )
    db.add(req)
    db.flush()

    for item in result.get("items", []):
        db.add(RequestItem(
            request_id=req.id,
            account_id=item.get("account_id"),
            field_name=item.get("field_name"),
            current_value=item.get("current_value"),
            proposed_value=item.get("proposed_value"),
        ))

    db.commit()
    db.refresh(req)

    return {
        "request_id": str(req.id),
        "classification_status": req.classification_status,
        "confidence": float(req.confidence) if req.confidence else None,
        "request_type": req.request_type,
        "vendor_number": req.vendor_number,
        "vendor_name": req.vendor_name,
        "items_count": len(req.items),
        "notes": req.notes,
    }